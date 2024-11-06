from fastapi import FastAPI, Form
from openpyxl import Workbook, load_workbook
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
from telegram import Bot
import os

app = FastAPI()
templates = Jinja2Templates(directory="templates")

# إعدادات ملف Excel
file_path = os.path.join(os.path.dirname(__file__), "data.xlsx")

# إعدادات تلغرام
TELEGRAM_API_TOKEN = '7731118993:AAE8-1Tc3xnjCvPOFUt59ldiK-4jnX888h0'
CHAT_ID = '6244988564'

def save_data_to_excel(name, email):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "Name"
        sheet['B1'] = "Email"

    next_row = sheet.max_row + 1
    sheet[f'A{next_row}'] = name
    sheet[f'B{next_row}'] = email
    workbook.save(file_path)

async def send_telegram_message(name: str, email: str):
    bot = Bot(token=TELEGRAM_API_TOKEN)
    message = f"تم استلام معلومات جديدة:\nالاسم: {name}\nالبريد الإلكتروني: {email}"
    await bot.send_message(chat_id=CHAT_ID, text=message)

async def send_excel_file_via_telegram():
    bot = Bot(token=TELEGRAM_API_TOKEN)
    with open(file_path, 'rb') as file:
        await bot.send_document(chat_id=CHAT_ID, document=file)

@app.get("/", response_class=HTMLResponse)
async def read_form(request: Request):
    return templates.TemplateResponse("form.html", {"request": request})

@app.post("/submit/")
async def handle_form(name: str = Form(...), email: str = Form(...)):
    save_data_to_excel(name, email)
    await send_telegram_message(name, email)
    return {"message": "تم حفظ البيانات بنجاح"}

@app.get("/send_excel/", response_class=HTMLResponse)
async def send_excel(request: Request):
    return templates.TemplateResponse("sendbutton.html", {"request": request})

@app.post("/send_excel_file/")
async def send_excel_file():
    await send_excel_file_via_telegram()
    return {"message": "تم إرسال ملف Excel بنجاح"}
